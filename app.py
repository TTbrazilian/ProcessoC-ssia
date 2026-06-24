"""
iG2P · Formulário de Inscrição — Processo Seletivo de Cássia - MG (Médicos)
Mesma estrutura do app de Ibiraci: tema iG2P, janela de datas, gravação em
planilha (Google Sheets + fallback Excel), e-mail de confirmação.
Diferenças desta versão: campos médicos (CRM, telefone) e ANEXOS de documentos,
enviados ao Google Drive (com fallback local) e linkados na planilha.
"""

import os
import re
import io
import ssl
import base64
import random
import smtplib
import unicodedata
from datetime import datetime, timezone, timedelta
from email.message import EmailMessage

import streamlit as st

APP_DIR = os.path.dirname(os.path.abspath(__file__))
PLANILHA_LOCAL = os.path.join(APP_DIR, "respostas_cassia.xlsx")
BANDEIRA_PATH = os.path.join(APP_DIR, "cássiabandeira.png")

# ----------------------------------------------------------------------------
# JANELA DE INSCRIÇÕES  ——  edite apenas estas duas datas para abrir/fechar.
# Fuso de Brasília (UTC-3). Use None para deixar BLOQUEADO ("em breve").
#   Exemplo para abrir:
#     DATA_ABERTURA     = datetime(2026, 7, 1,  8, 0, tzinfo=TZ)
#     DATA_ENCERRAMENTO = datetime(2026, 7, 31, 23, 59, tzinfo=TZ)
# ----------------------------------------------------------------------------
TZ = timezone(timedelta(hours=-3))
# ABERTO PARA TESTES — antes de divulgar, troque pelas datas reais (ou volte ambas para None p/ bloquear).
DATA_ABERTURA = datetime(2020, 1, 1, 0, 0, tzinfo=TZ)
DATA_ENCERRAMENTO = None

# Anexos
EXTS_PERMITIDAS = ["pdf", "jpg", "jpeg", "png"]
TAMANHO_MAX_MB = 10

# Definição dos 4 anexos: (chave, rótulo)
ANEXOS = [
    ("diploma", "Diploma da Graduação em Medicina"),
    ("pos", "Certificado de Pós-Graduação ou Especialidades"),
    ("aps", "Comprovante de Tempo de Serviço na APS"),
    ("cassia", "Comprovante de Tempo de Serviço no Município de Cássia"),
]

COLUNAS_PLANILHA = [
    "Data/Hora do Envio", "Protocolo", "Nome Completo", "Data de Nascimento",
    "CPF", "CRM", "Telefone", "E-mail",
    "Diploma Graduação Medicina", "Certificado Pós/Especialidade",
    "Tempo de Serviço na APS", "Tempo de Serviço em Cássia",
]

# ----------------------------------------------------------------------------
# Tema (padrão iG2P)
# ----------------------------------------------------------------------------
TEMAS = {
    "claro": {
        "bg": "#F4F5FB", "card": "#FFFFFF", "texto": "#1E1F36",
        "texto2": "#5A5C77", "borda": "#E3E5F0", "primaria": "#5B5BD6",
        "primaria_grad": "#7C5CFF", "faixa_txt": "#FFFFFF", "input_bg": "#FFFFFF",
        "sombra": "0 8px 24px rgba(91,91,214,0.08)", "ok": "#1FA971", "erro": "#E5484D",
    },
    "escuro": {
        "bg": "#0F1020", "card": "#1A1B33", "texto": "#ECEDF7",
        "texto2": "#9FA1C0", "borda": "#2C2E4A", "primaria": "#7C7CFF",
        "primaria_grad": "#9D7BFF", "faixa_txt": "#FFFFFF", "input_bg": "#23244180",
        "sombra": "0 8px 24px rgba(0,0,0,0.35)", "ok": "#3DD68C", "erro": "#FF6369",
    },
}

st.set_page_config(
    page_title="Inscrição · Processo Seletivo Cássia",
    page_icon="🏥",
    layout="centered",
)

for k, v in {"modo": "claro", "enviado": False, "protocolo": None, "destino": None,
             "motivo": None, "email_ok": None, "email_dest": None,
             "anexos_local": False}.items():
    if k not in st.session_state:
        st.session_state[k] = v

T = TEMAS[st.session_state.modo]


# ----------------------------------------------------------------------------
# CSS — injeta o tema iG2P sobre os componentes do Streamlit
# ----------------------------------------------------------------------------
def aplicar_css(t):
    st.markdown(f"""
    <style>
    .stApp {{ background:{t['bg']}; }}
    #MainMenu, footer, header {{ visibility:hidden; }}
    .block-container {{ padding-top:1.2rem; padding-bottom:3rem; max-width:820px; }}

    .ig2p-header {{
        background:linear-gradient(135deg,{t['primaria']} 0%,{t['primaria_grad']} 100%);
        border-radius:20px; padding:26px 30px; color:#fff;
        box-shadow:{t['sombra']}; display:flex; align-items:center; gap:20px;
        margin-bottom:6px;
    }}
    .ig2p-flag {{
        width:62px; height:62px; border-radius:14px; flex:0 0 auto;
        background:rgba(255,255,255,.18); display:flex; align-items:center;
        justify-content:center; font-weight:800; font-size:20px; letter-spacing:1px;
        border:2px solid rgba(255,255,255,.35); overflow:hidden;
    }}
    .ig2p-flag img {{ width:100%; height:100%; object-fit:cover; }}
    .ig2p-header h1 {{ font-size:1.42rem; margin:0; line-height:1.18; font-weight:800; }}
    .ig2p-header p {{ margin:4px 0 0; opacity:.92; font-size:.86rem; }}
    .ig2p-chip {{
        display:inline-block; margin-top:8px; padding:3px 12px; border-radius:30px;
        background:rgba(255,255,255,.20); font-size:.72rem; font-weight:600;
    }}

    .ig2p-intro {{
        background:{t['card']}; color:{t['texto2']}; border:1px solid {t['borda']};
        border-radius:16px; padding:16px 20px; font-size:.84rem; line-height:1.5;
        margin:14px 0 4px; box-shadow:{t['sombra']};
    }}
    .ig2p-intro b {{ color:{t['texto']}; }}

    .ig2p-faixa {{
        background:linear-gradient(135deg,{t['primaria']} 0%,{t['primaria_grad']} 100%);
        color:{t['faixa_txt']}; padding:12px 18px; border-radius:12px;
        font-weight:700; font-size:1rem; margin:26px 0 10px;
        box-shadow:{t['sombra']};
    }}
    .ig2p-legal {{
        background:{t['card']}; border:1px dashed {t['borda']}; border-radius:12px;
        padding:14px 18px; font-size:.78rem; color:{t['texto2']}; line-height:1.55;
        margin-bottom:6px;
    }}
    .ig2p-legal b {{ color:{t['texto']}; }}

    label, .stMarkdown, .stRadio, p {{ color:{t['texto']} !important; }}
    .stRadio label p {{ color:{t['texto']} !important; }}

    .stTextInput input, .stDateInput input, .stTextArea textarea {{
        background:{t['input_bg']} !important; color:{t['texto']} !important;
        border:1px solid {t['borda']} !important; border-radius:10px !important;
    }}
    .stTextInput input:focus, .stTextArea textarea:focus {{
        border-color:{t['primaria']} !important;
        box-shadow:0 0 0 3px {t['primaria']}33 !important;
    }}

    .stButton > button {{
        background:linear-gradient(135deg,{t['primaria']} 0%,{t['primaria_grad']} 100%);
        color:#fff !important; border:none; border-radius:12px; padding:12px 28px;
        font-weight:700; font-size:.98rem; width:100%; transition:transform .15s ease;
        box-shadow:{t['sombra']};
    }}
    .stButton > button:hover {{ transform:translateY(-2px); }}

    /* File uploader no tema */
    [data-testid="stFileUploaderDropzone"] {{
        background:{t['input_bg']} !important; border:1px dashed {t['borda']} !important;
        border-radius:12px !important;
    }}

    .ig2p-card {{
        background:{t['card']}; border:1px solid {t['borda']}; border-radius:16px;
        padding:22px 24px; margin-top:10px; box-shadow:{t['sombra']};
    }}
    .ig2p-success {{
        background:{t['card']}; border:1px solid {t['ok']}; border-radius:18px;
        padding:30px; text-align:center; box-shadow:{t['sombra']};
    }}
    .ig2p-success .check {{
        width:64px; height:64px; border-radius:50%; background:{t['ok']}22;
        color:{t['ok']}; display:flex; align-items:center; justify-content:center;
        font-size:34px; margin:0 auto 14px;
    }}
    .ig2p-prot {{
        font-family:ui-monospace,monospace; font-size:1.4rem; font-weight:800;
        color:{t['primaria']}; letter-spacing:2px; margin:6px 0 2px;
    }}
    .ig2p-rodape {{ text-align:center; color:{t['texto2']}; font-size:.72rem; margin-top:26px; }}
    </style>
    """, unsafe_allow_html=True)


aplicar_css(T)


def flag_html():
    if os.path.exists(BANDEIRA_PATH):
        b64 = base64.b64encode(open(BANDEIRA_PATH, "rb").read()).decode()
        return f'<div class="ig2p-flag"><img src="data:image/png;base64,{b64}"/></div>'
    return '<div class="ig2p-flag">CA</div>'


# ----------------------------------------------------------------------------
# Validações
# ----------------------------------------------------------------------------
def so_digitos(v):
    return re.sub(r"\D", "", v or "")


def email_valido(e):
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", (e or "").strip()))


def cpf_valido(cpf):
    cpf = so_digitos(cpf)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False
    for i in (9, 10):
        soma = sum(int(cpf[n]) * ((i + 1) - n) for n in range(i))
        dig = (soma * 10) % 11
        dig = 0 if dig == 10 else dig
        if dig != int(cpf[i]):
            return False
    return True


def data_valida(s):
    try:
        datetime.strptime((s or "").strip(), "%d/%m/%Y")
        return True
    except ValueError:
        return False


def telefone_valido(t):
    return len(so_digitos(t)) in (10, 11)


def gerar_protocolo():
    return f"CAS-{datetime.now():%Y%m%d}-{random.randint(1000, 9999)}"


def slug(texto):
    t = unicodedata.normalize("NFKD", texto or "").encode("ascii", "ignore").decode()
    t = re.sub(r"[^A-Za-z0-9]+", "_", t).strip("_")
    return t[:40] or "candidato"


# ----------------------------------------------------------------------------
# Máscaras
# ----------------------------------------------------------------------------
def mascara_cpf(texto):
    d = so_digitos(texto)[:11]
    if len(d) > 9:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    if len(d) > 6:
        return f"{d[:3]}.{d[3:6]}.{d[6:]}"
    if len(d) > 3:
        return f"{d[:3]}.{d[3:]}"
    return d


def mascara_data(texto):
    d = so_digitos(texto)[:8]
    if len(d) > 4:
        return f"{d[:2]}/{d[2:4]}/{d[4:]}"
    if len(d) > 2:
        return f"{d[:2]}/{d[2:]}"
    return d


def mascara_telefone(texto):
    d = so_digitos(texto)[:11]
    if len(d) >= 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:]}"
    if len(d) >= 7:
        return f"({d[:2]}) {d[2:6]}-{d[6:]}"
    if len(d) >= 3:
        return f"({d[:2]}) {d[2:]}"
    return d


def fmt_cpf():
    st.session_state.cpf = mascara_cpf(st.session_state.cpf)


def fmt_nascimento():
    st.session_state.nascimento = mascara_data(st.session_state.nascimento)


def fmt_telefone():
    st.session_state.telefone = mascara_telefone(st.session_state.telefone)


# ----------------------------------------------------------------------------
# Janela de inscrições
# ----------------------------------------------------------------------------
def agora():
    return datetime.now(TZ)


def status_inscricoes():
    if DATA_ABERTURA is None:
        return "antes"
    n = agora()
    if n < DATA_ABERTURA:
        return "antes"
    if DATA_ENCERRAMENTO is not None and n > DATA_ENCERRAMENTO:
        return "encerrado"
    return "aberto"


def fmt_data_br(dt):
    return dt.strftime("%d/%m/%Y às %Hh%M") if dt else "a definir"


# ----------------------------------------------------------------------------
# E-mail de confirmação (SMTP via st.secrets[email])
# ----------------------------------------------------------------------------
def enviar_email_confirmacao(destinatario, nome, protocolo):
    if "email" not in st.secrets:
        return False, "Seção [email] não configurada nos secrets."
    cfg = st.secrets["email"]
    try:
        remetente = cfg["remetente"]              # endereço que aparece no "De:"
        usuario = cfg.get("usuario", remetente)   # login SMTP (pode diferir do remetente)
        senha = cfg["senha"]
        host = cfg.get("host", "smtp.gmail.com")
        porta = int(cfg.get("porta", 465))
        nome_exib = cfg.get("nome_exibicao", "Processo Seletivo Cássia")
    except Exception as e:
        return False, f"Configuração de e-mail incompleta: {e}"

    assunto = "Confirmação de Inscrição — Processo Seletivo Cássia nº 001/2026"
    quando = agora().strftime("%d/%m/%Y às %Hh%M")
    texto = (
        f"Olá, {nome}!\n\n"
        f"Sua inscrição no Processo Seletivo nº 001/2026 do Município de Cássia - MG "
        f"foi recebida com sucesso.\n\n"
        f"Número de protocolo: {protocolo}\n"
        f"Data/hora: {quando}\n\n"
        f"Guarde este número de protocolo — ele é o comprovante da sua inscrição.\n\n"
        f"As informações e documentos enviados são de inteira responsabilidade do candidato.\n\n"
        f"Este é um e-mail automático, não responda.\n"
        f"— {nome_exib}"
    )
    html = f"""\
<div style="font-family:Arial,sans-serif;max-width:560px;margin:auto;border:1px solid #E3E5F0;border-radius:14px;overflow:hidden">
  <div style="background:linear-gradient(135deg,#5B5BD6,#7C5CFF);color:#fff;padding:22px 26px">
    <h2 style="margin:0;font-size:18px">Inscrição confirmada ✓</h2>
    <p style="margin:4px 0 0;opacity:.9;font-size:13px">Processo Seletivo nº 001/2026 · Cássia - MG</p>
  </div>
  <div style="padding:24px 26px;color:#1E1F36;font-size:14px;line-height:1.6">
    <p>Olá, <b>{nome}</b>! Sua inscrição foi recebida com sucesso.</p>
    <p style="margin:18px 0;text-align:center">
      <span style="display:inline-block;background:#5B5BD61a;color:#5B5BD6;font-weight:800;
      font-size:20px;letter-spacing:2px;padding:12px 22px;border-radius:10px">{protocolo}</span>
    </p>
    <p style="margin:6px 0"><b>Data/hora:</b> {quando}</p>
    <p style="margin-top:16px;color:#5A5C77;font-size:13px">
      Guarde este número de protocolo — ele é o comprovante da sua inscrição.
      As informações e documentos enviados são de inteira responsabilidade do candidato.</p>
    <p style="color:#9FA1C0;font-size:12px;margin-top:18px">Este é um e-mail automático, não responda.</p>
  </div>
</div>"""

    msg = EmailMessage()
    msg["Subject"] = assunto
    msg["From"] = f"{nome_exib} <{remetente}>"
    msg["To"] = destinatario
    msg.set_content(texto)
    msg.add_alternative(html, subtype="html")
    try:
        ctx = ssl.create_default_context()
        if porta == 465:
            with smtplib.SMTP_SSL(host, porta, context=ctx, timeout=20) as s:
                s.login(usuario, senha)
                s.send_message(msg)
        else:
            with smtplib.SMTP(host, porta, timeout=20) as s:
                s.starttls(context=ctx)
                s.login(usuario, senha)
                s.send_message(msg)
        return True, None
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


# ----------------------------------------------------------------------------
# Anexos — Google Drive (prioridade) ou pasta local (fallback)
# ----------------------------------------------------------------------------
def _drive_service():
    from google.oauth2.service_account import Credentials  # type: ignore
    from googleapiclient.discovery import build  # type: ignore
    escopos = ["https://www.googleapis.com/auth/drive"]
    cred = Credentials.from_service_account_info(
        dict(st.secrets["gcp_service_account"]), scopes=escopos)
    return build("drive", "v3", credentials=cred, cache_discovery=False)


def upload_arquivo_drive(uploaded, nome_final, folder_id):
    from googleapiclient.http import MediaIoBaseUpload  # type: ignore
    service = _drive_service()
    media = MediaIoBaseUpload(
        io.BytesIO(uploaded.getvalue()),
        mimetype=uploaded.type or "application/octet-stream", resumable=False)
    meta = {"name": nome_final, "parents": [folder_id]}
    arq = service.files().create(
        body=meta, media_body=media, fields="id, webViewLink",
        supportsAllDrives=True).execute()
    comissao = st.secrets.get("drive", {}).get("compartilhar_email", "")
    if comissao:
        try:
            service.permissions().create(
                fileId=arq["id"],
                body={"type": "user", "role": "reader", "emailAddress": comissao},
                sendNotificationEmail=False, supportsAllDrives=True).execute()
        except Exception:
            pass
    return arq.get("webViewLink", arq.get("id", ""))


def salvar_arquivo(uploaded, nome_base):
    """Retorna (link_ou_caminho, destino, motivo). destino: 'drive' | 'local' | 'erro'."""
    ext = os.path.splitext(uploaded.name)[1].lower()
    nome_final = f"{nome_base}{ext}"
    folder_id = st.secrets.get("drive", {}).get("folder_id", "")
    motivo_drive = ""
    if "gcp_service_account" in st.secrets and folder_id:
        try:
            return upload_arquivo_drive(uploaded, nome_final, folder_id), "drive", None
        except Exception as e:
            motivo_drive = f"{type(e).__name__}: {e}"
    else:
        motivo_drive = "Drive não configurado ([drive].folder_id ausente)."
    # fallback local
    try:
        pasta = os.path.join(APP_DIR, "anexos")
        os.makedirs(pasta, exist_ok=True)
        caminho = os.path.join(pasta, nome_final)
        with open(caminho, "wb") as fp:
            fp.write(uploaded.getvalue())
        return caminho, "local", motivo_drive
    except Exception as e:
        return None, "erro", f"Drive: {motivo_drive} | Local: {type(e).__name__}: {e}"


# ----------------------------------------------------------------------------
# Planilha — Google Sheets (prioridade) ou Excel local (fallback)
# ----------------------------------------------------------------------------
def conectar_sheets():
    if "gcp_service_account" not in st.secrets:
        return None, ("Secrets não foram lidos. No Streamlit Cloud: cole o conteúdo do "
                      "secrets.toml em Manage app → Settings → Secrets. Localmente: confirme "
                      "que o arquivo está em .streamlit/secrets.toml e rode o app da pasta do projeto.")
    sheet_id = st.secrets.get("planilha", {}).get("sheet_id", "")
    if not sheet_id or "COLE_AQUI" in sheet_id:
        return None, "sheet_id não configurado em [planilha] no secrets.toml."
    try:
        import gspread  # type: ignore
        from google.oauth2.service_account import Credentials  # type: ignore
    except ImportError:
        return None, ("Bibliotecas gspread/google-auth NÃO instaladas. "
                      "Rode: py -m pip install gspread google-auth")
    try:
        escopos = ["https://www.googleapis.com/auth/spreadsheets"]
        cred = Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=escopos)
        gc = gspread.authorize(cred)
        sh = gc.open_by_key(sheet_id)
        return sh.sheet1, None
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


def salvar_google_sheets(linha):
    ws, motivo = conectar_sheets()
    if ws is None:
        return False, motivo
    try:
        if not ws.row_values(1):
            ws.append_row(COLUNAS_PLANILHA, value_input_option="USER_ENTERED")
        ws.append_row(linha, value_input_option="USER_ENTERED")
        return True, None
    except Exception as e:
        return False, f"Falha ao gravar a linha: {type(e).__name__}: {e}"


def salvar_excel_local(linha):
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        return False, "openpyxl não está instalado neste ambiente."
    try:
        if os.path.exists(PLANILHA_LOCAL):
            wb = load_workbook(PLANILHA_LOCAL)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Inscrições"
            ws.append(COLUNAS_PLANILHA)
        ws.append(linha)
        r = ws.max_row
        for col in (2, 5, 6, 7):  # Protocolo, CPF, CRM, Telefone como texto
            ws.cell(row=r, column=col).number_format = "@"
        wb.save(PLANILHA_LOCAL)
        return True, None
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def salvar_resposta(linha):
    ok, motivo = salvar_google_sheets(linha)
    if ok:
        return "google", None
    ok_local, _ = salvar_excel_local(linha)
    if ok_local:
        return "local", motivo
    return "erro", motivo


# ----------------------------------------------------------------------------
# Tela de sucesso
# ----------------------------------------------------------------------------
if st.session_state.enviado:
    if st.session_state.destino == "google":
        badge = (f'<div style="margin-top:14px;padding:8px 14px;border-radius:10px;'
                 f'background:{T["ok"]}1a;color:{T["ok"]};font-size:.8rem;font-weight:600;">'
                 f'✓ Registrado no Google Sheets</div>')
    else:
        badge = (f'<div style="margin-top:14px;padding:10px 14px;border-radius:10px;'
                 f'background:#FFB02022;color:#B7791F;font-size:.78rem;text-align:left;">'
                 f'<b>Atenção:</b> registrado apenas na planilha local, não no Google Sheets.<br>'
                 f'<b>Motivo:</b> {st.session_state.motivo}</div>')
    if st.session_state.anexos_local:
        badge += (f'<div style="margin-top:8px;padding:10px 14px;border-radius:10px;'
                  f'background:#FFB02022;color:#B7791F;font-size:.78rem;text-align:left;">'
                  f'<b>Atenção:</b> os anexos foram salvos apenas localmente (Drive não '
                  f'configurado). Em produção, configure o Google Drive.</div>')
    if st.session_state.email_ok is True:
        email_badge = (f'<p style="color:{T["texto2"]};font-size:.78rem;margin-top:10px;">'
                       f'📧 Um e-mail de confirmação foi enviado para '
                       f'<b>{st.session_state.email_dest}</b>.</p>')
    elif st.session_state.email_ok is False:
        email_badge = (f'<p style="color:{T["texto2"]};font-size:.78rem;margin-top:10px;">'
                       f'Não foi possível enviar o e-mail de confirmação agora, mas sua '
                       f'inscrição está registrada. Guarde o protocolo acima.</p>')
    else:
        email_badge = ""
    st.markdown(f"""
    <div class="ig2p-success">
        <div class="check">✓</div>
        <h2 style="color:{T['texto']};margin:0;">Inscrição enviada com sucesso!</h2>
        <p style="color:{T['texto2']};margin:8px 0 0;">Guarde o número de protocolo abaixo:</p>
        <div class="ig2p-prot">{st.session_state.protocolo}</div>
        <p style="color:{T['texto2']};font-size:.8rem;margin-top:14px;">
            Suas informações e documentos foram registrados. A constatação de falsidade ou
            omissão de dados implicará na eliminação do candidato.
        </p>
        {email_badge}
        {badge}
    </div>
    """, unsafe_allow_html=True)
    if st.button("Nova inscrição"):
        for k in ("enviado", "protocolo", "anexos_local"):
            st.session_state[k] = False if k == "enviado" or k == "anexos_local" else None
        st.rerun()
    st.markdown('<div class="ig2p-rodape">iG2P · Inteligência em Gestão Pública · G3ST4O · Prefeitura de Cássia - MG</div>',
                unsafe_allow_html=True)
    st.stop()


# ----------------------------------------------------------------------------
# Bloqueio por janela de datas
# ----------------------------------------------------------------------------
_status = status_inscricoes()
if _status != "aberto":
    if _status == "antes":
        icone, titulo = "🔒", "Inscrições ainda não abertas"
        msg = ("O período de inscrições será divulgado em breve." if DATA_ABERTURA is None
               else f"As inscrições abrem em <b>{fmt_data_br(DATA_ABERTURA)}</b>.")
    else:
        icone, titulo = "⛔", "Inscrições encerradas"
        msg = f"O período de inscrições foi encerrado em <b>{fmt_data_br(DATA_ENCERRAMENTO)}</b>."
    periodo = ""
    if DATA_ABERTURA is not None:
        periodo = (f'<div style="margin-top:16px;padding:12px 16px;border-radius:12px;'
                   f'background:{T["primaria"]}14;color:{T["texto2"]};font-size:.82rem;">'
                   f'<b style="color:{T["texto"]};">Período de inscrições</b><br>'
                   f'Abertura: {fmt_data_br(DATA_ABERTURA)}<br>'
                   f'Encerramento: {fmt_data_br(DATA_ENCERRAMENTO)}</div>')
    st.markdown(f"""
    <div class="ig2p-header" style="margin-bottom:18px;">
        {flag_html()}
        <div><h1>Processo Seletivo nº 001/2026</h1>
        <p>Município de Cássia · Estado de Minas Gerais</p></div>
    </div>
    <div class="ig2p-success">
        <div class="check" style="background:{T['primaria']}1a;color:{T['primaria']};">{icone}</div>
        <h2 style="color:{T['texto']};margin:0;">{titulo}</h2>
        <p style="color:{T['texto2']};margin:10px 0 0;font-size:.9rem;">{msg}</p>
        {periodo}
    </div>
    <div class="ig2p-rodape">iG2P · Inteligência em Gestão Pública · G3ST4O · Prefeitura de Cássia - MG</div>
    """, unsafe_allow_html=True)
    st.stop()


# ----------------------------------------------------------------------------
# Cabeçalho + alternância de tema
# ----------------------------------------------------------------------------
col_a, col_b = st.columns([6, 1])
with col_b:
    if st.button("🌙" if st.session_state.modo == "claro" else "☀️", help="Alternar tema"):
        st.session_state.modo = "escuro" if st.session_state.modo == "claro" else "claro"
        st.rerun()

st.markdown(f"""
<div class="ig2p-header">
    {flag_html()}
    <div>
        <h1>Formulário de Inscrição — Processo Seletivo nº 001/2026</h1>
        <p>Município de Cássia · Estado de Minas Gerais</p>
        <span class="ig2p-chip">Seleção de profissionais médicos</span>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="ig2p-intro">
O Município de Cássia, Estado de Minas Gerais, realizará processo seletivo para
contratação de profissionais médicos para atender necessidade de excepcional interesse
público.<br><br>
<b>As informações e documentos enviados são de inteira responsabilidade do candidato.</b>
A constatação, a qualquer tempo, de falsidade ou omissão implicará na eliminação do
candidato e na anulação de sua inscrição.<br><br>
<span style="color:{T['erro']};font-weight:600;">*</span> Indica um campo obrigatório.
</div>
""", unsafe_allow_html=True)

# ----------------------------------------------------------------------------
# DADOS DO CANDIDATO
# ----------------------------------------------------------------------------
st.markdown('<div class="ig2p-faixa">Dados do Candidato</div>', unsafe_allow_html=True)

nome = st.text_input("1. Nome Completo (sem abreviações) *")
nascimento = st.text_input("2. Data de Nascimento (dia/mês/ano) — Ex: 01/01/1990 *",
                           key="nascimento", max_chars=10, on_change=fmt_nascimento,
                           help="Digite só os números — a barra é inserida automaticamente.")
cpf = st.text_input("3. CPF — digite apenas os números *",
                    key="cpf", max_chars=14, on_change=fmt_cpf,
                    help="A pontuação (000.000.000-00) é aplicada automaticamente.")
crm = st.text_input("4. CRM (número de registro no Conselho Regional de Medicina) *",
                    max_chars=20, help="Ex: 12345/MG")
telefone = st.text_input("5. Telefone / WhatsApp *",
                         key="telefone", max_chars=16, on_change=fmt_telefone,
                         help="Digite só os números — o formato (00) 00000-0000 é aplicado automaticamente.")
email = st.text_input("6. E-mail *")

# ----------------------------------------------------------------------------
# DOCUMENTOS (ANEXOS)
# ----------------------------------------------------------------------------
st.markdown('<div class="ig2p-faixa">Documentos (anexos)</div>', unsafe_allow_html=True)
st.markdown(f"""
<div class="ig2p-legal">
Anexe os documentos abaixo nos formatos <b>{', '.join(EXTS_PERMITIDAS).upper()}</b>
(até {TAMANHO_MAX_MB} MB cada). Todos são obrigatórios.
</div>
""", unsafe_allow_html=True)

arquivos = {}
for i, (chave, rotulo) in enumerate(ANEXOS, start=7):
    arquivos[chave] = st.file_uploader(f"{i}. {rotulo} *", type=EXTS_PERMITIDAS, key=f"file_{chave}")

# ----------------------------------------------------------------------------
# ENVIO + VALIDAÇÃO
# ----------------------------------------------------------------------------
st.markdown("<br>", unsafe_allow_html=True)
if st.button("Enviar inscrição"):
    erros = []
    if not nome.strip() or len(nome.split()) < 2:
        erros.append("Informe o nome completo, sem abreviações.")
    if not data_valida(nascimento):
        erros.append("Data de nascimento inválida — use o formato dia/mês/ano.")
    if not cpf_valido(cpf):
        erros.append("CPF inválido — informe os 11 números.")
    if not crm.strip():
        erros.append("Informe o número do CRM.")
    if not telefone_valido(telefone):
        erros.append("Telefone inválido — informe DDD + número.")
    if not email_valido(email):
        erros.append("Informe um e-mail válido.")
    for chave, rotulo in ANEXOS:
        arq = arquivos[chave]
        if arq is None:
            erros.append(f"Anexe: {rotulo}.")
        elif len(arq.getvalue()) > TAMANHO_MAX_MB * 1024 * 1024:
            erros.append(f"O arquivo '{rotulo}' excede {TAMANHO_MAX_MB} MB.")

    if erros:
        st.markdown(
            f'<div class="ig2p-card" style="border-color:{T["erro"]};">'
            f'<b style="color:{T["erro"]};">Corrija os campos abaixo:</b><ul>'
            + "".join(f"<li style=\"color:{T['texto2']};\">{e}</li>" for e in erros)
            + "</ul></div>", unsafe_allow_html=True)
    else:
        protocolo = gerar_protocolo()
        with st.spinner("Enviando documentos e registrando a inscrição..."):
            links = {}
            anexos_local = False
            falha_anexo = None
            for chave, rotulo in ANEXOS:
                arq = arquivos[chave]
                base = f"{protocolo}_{chave}_{slug(nome)}"
                link, dest, motivo = salvar_arquivo(arq, base)
                if dest == "erro":
                    falha_anexo = f"{rotulo}: {motivo}"
                    break
                if dest == "local":
                    anexos_local = True
                links[chave] = link

            if falha_anexo:
                st.markdown(
                    f'<div class="ig2p-card" style="border-color:{T["erro"]};">'
                    f'<b style="color:{T["erro"]};">Não foi possível enviar seus documentos.</b><br>'
                    f'<span style="color:{T["texto2"]};font-size:.85rem;">Tente novamente em '
                    f'instantes. Se persistir, avise a organização.<br>'
                    f'<b>Detalhe:</b> {falha_anexo}</span></div>', unsafe_allow_html=True)
            else:
                linha = [
                    agora().strftime("%d/%m/%Y %H:%M:%S"), protocolo,
                    nome.strip(), nascimento.strip(), cpf.strip(), crm.strip(),
                    telefone.strip(), email.strip(),
                    links.get("diploma", ""), links.get("pos", ""),
                    links.get("aps", ""), links.get("cassia", ""),
                ]
                destino, motivo = salvar_resposta(linha)
                if destino == "erro":
                    st.markdown(
                        f'<div class="ig2p-card" style="border-color:{T["erro"]};">'
                        f'<b style="color:{T["erro"]};">Não foi possível registrar sua inscrição.</b><br>'
                        f'<span style="color:{T["texto2"]};font-size:.85rem;">Tente novamente em '
                        f'instantes.<br><b>Detalhe técnico:</b> {motivo}</span></div>',
                        unsafe_allow_html=True)
                else:
                    email_ok, _ = enviar_email_confirmacao(email.strip(), nome.strip(), protocolo)
                    st.session_state.email_ok = email_ok
                    st.session_state.email_dest = email.strip()
                    st.session_state.anexos_local = anexos_local
                    st.session_state.destino = destino
                    st.session_state.motivo = motivo
                    st.session_state.protocolo = protocolo
                    st.session_state.enviado = True
                    st.rerun()

st.markdown('<div class="G3ST40-rodape"> G3ST4O · Prefeitura de Cássia - MG</div>',
            unsafe_allow_html=True)